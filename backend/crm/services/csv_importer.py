"""
Contact Importer Service

Handles bulk import of contacts from CSV and Excel files with:
- Flexible column mapping
- Duplicate detection
- Error reporting
- Optional deal creation

Supported formats:
- CSV (.csv)
- Excel (.xlsx, .xls)
"""

import csv
import io
import logging
from typing import Dict, List, Optional, Tuple
from django.utils import timezone
from django.db import transaction

logger = logging.getLogger(__name__)


def parse_file_to_rows(file_obj, filename: str) -> Tuple[List[str], List[Dict]]:
    """
    Parse a file (CSV or Excel) into rows.

    Args:
        file_obj: File object or file content
        filename: Original filename (used to detect format)

    Returns:
        Tuple of (headers, rows as list of dicts)
    """
    filename_lower = filename.lower()

    # Read file content
    if hasattr(file_obj, 'read'):
        content = file_obj.read()
        if hasattr(file_obj, 'seek'):
            file_obj.seek(0)
    else:
        content = file_obj

    # Excel formats
    if filename_lower.endswith(('.xlsx', '.xls')):
        return _parse_excel(content)

    # Default: CSV
    return _parse_csv(content)


def _parse_csv(content: bytes) -> Tuple[List[str], List[Dict]]:
    """Parse CSV content."""
    if isinstance(content, bytes):
        content = content.decode('utf-8-sig')

    reader = csv.DictReader(io.StringIO(content))
    headers = reader.fieldnames or []
    rows = [dict(row) for row in reader]

    return headers, rows


def _parse_excel(content: bytes) -> Tuple[List[str], List[Dict]]:
    """Parse Excel content using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel support. Install with: pip install openpyxl")

    # Load workbook from bytes
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows_data = list(ws.iter_rows(values_only=True))

    if not rows_data:
        return [], []

    # First row is headers
    headers = [str(h).strip() if h else f'Column_{i}' for i, h in enumerate(rows_data[0])]

    # Rest are data rows
    rows = []
    for row in rows_data[1:]:
        if not any(row):  # Skip empty rows
            continue
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers):
                # Convert to string, handle None
                row_dict[headers[i]] = str(value).strip() if value is not None else ''
        rows.append(row_dict)

    wb.close()
    return headers, rows

# Standard column mappings (CSV header -> model field)
COLUMN_MAPPINGS = {
    # Email variations
    'email': 'email',
    'email_address': 'email',
    'e-mail': 'email',
    'contact_email': 'email',

    # Name variations
    'name': 'name',
    'full_name': 'name',
    'contact_name': 'name',
    'contact': 'name',
    'first_name': 'first_name',  # Will be combined
    'last_name': 'last_name',    # Will be combined
    'firstname': 'first_name',
    'lastname': 'last_name',

    # Company variations
    'company': 'company',
    'company_name': 'company',
    'organization': 'company',
    'business': 'company',
    'business_name': 'company',

    # Website variations
    'website': 'website',
    'website_url': 'website',
    'url': 'website',
    'site': 'website',
    'domain': 'website',

    # Domain Authority
    'domain_authority': 'domain_authority',
    'da': 'domain_authority',
    'dr': 'domain_authority',  # Ahrefs Domain Rating
    'authority': 'domain_authority',

    # Notes
    'notes': 'notes',
    'note': 'notes',
    'comments': 'notes',
    'description': 'notes',

    # Tags
    'tags': 'tags',
    'tag': 'tags',
    'labels': 'tags',
    'categories': 'tags',
}


class ContactImporter:
    """Import contacts from CSV or Excel files."""

    def __init__(self, contact_import):
        """
        Initialize importer with a ContactImport instance.

        Args:
            contact_import: ContactImport model instance
        """
        self.contact_import = contact_import
        self.brand = contact_import.brand
        self.contact_type = contact_import.contact_type
        self.source = contact_import.source
        self.create_deals = contact_import.create_deals
        self.pipeline = contact_import.pipeline

        self.imported = []
        self.skipped = []
        self.errors = []

    def process(self) -> Dict:
        """
        Process the file (CSV or Excel) and import contacts.

        Returns:
            Dict with import results
        """
        from crm.models import Contact, Deal, ContactImport

        self.contact_import.status = 'processing'
        self.contact_import.save()

        try:
            # Read and parse file (CSV or Excel)
            file_content = self.contact_import.file.read()
            filename = self.contact_import.file_name or self.contact_import.file.name
            headers, rows = parse_file_to_rows(file_content, filename)

            self.contact_import.total_rows = len(rows)
            self.contact_import.save()

            # Map columns
            column_map = self._map_columns(headers)

            if 'email' not in column_map.values():
                raise ValueError("File must have an 'email' column")

            # Process rows
            for row_num, row in enumerate(rows, start=2):  # Start at 2 (header is row 1)
                try:
                    result = self._process_row(row, column_map, row_num)
                    if result == 'imported':
                        self.imported.append(row_num)
                    elif result == 'skipped':
                        self.skipped.append(row_num)
                except Exception as e:
                    self.errors.append({
                        'row': row_num,
                        'error': str(e),
                        'data': dict(row)
                    })

            # Update import record
            self.contact_import.status = 'completed'
            self.contact_import.imported_count = len(self.imported)
            self.contact_import.skipped_count = len(self.skipped)
            self.contact_import.error_count = len(self.errors)
            self.contact_import.errors = self.errors[:100]  # Limit stored errors
            self.contact_import.completed_at = timezone.now()
            self.contact_import.save()

            logger.info(
                f"Import completed: {len(self.imported)} imported, "
                f"{len(self.skipped)} skipped, {len(self.errors)} errors"
            )

            return {
                'success': True,
                'imported': len(self.imported),
                'skipped': len(self.skipped),
                'errors': len(self.errors),
                'error_details': self.errors[:10]
            }

        except Exception as e:
            logger.error(f"Import failed: {e}")
            self.contact_import.status = 'failed'
            self.contact_import.errors = [{'error': str(e)}]
            self.contact_import.save()

            return {
                'success': False,
                'error': str(e)
            }

    def _map_columns(self, fieldnames: List[str]) -> Dict[str, str]:
        """
        Map CSV column names to model fields.

        Args:
            fieldnames: List of CSV column headers

        Returns:
            Dict mapping CSV columns to model fields
        """
        column_map = {}

        for col in fieldnames:
            col_lower = col.lower().strip()
            if col_lower in COLUMN_MAPPINGS:
                column_map[col] = COLUMN_MAPPINGS[col_lower]
            else:
                # Try partial matching
                for key, field in COLUMN_MAPPINGS.items():
                    if key in col_lower or col_lower in key:
                        column_map[col] = field
                        break

        return column_map

    def _process_row(self, row: Dict, column_map: Dict, row_num: int) -> str:
        """
        Process a single CSV row.

        Args:
            row: CSV row as dict
            column_map: Column to field mapping
            row_num: Row number for error reporting

        Returns:
            'imported', 'skipped', or raises exception
        """
        from crm.models import Contact, Deal

        # Extract mapped values
        data = {}
        for csv_col, model_field in column_map.items():
            value = row.get(csv_col, '').strip()
            if value:
                data[model_field] = value

        # Handle first_name + last_name combination
        if 'first_name' in data or 'last_name' in data:
            first = data.pop('first_name', '')
            last = data.pop('last_name', '')
            data['name'] = f"{first} {last}".strip()

        # Validate email
        email = data.get('email', '').lower()
        if not email or '@' not in email:
            raise ValueError(f"Invalid email: {email}")

        data['email'] = email

        # Check for duplicate within same brand (case-insensitive)
        existing = Contact.objects.filter(email__iexact=email, brand=self.brand).first()
        if existing:
            logger.debug(f"Skipping duplicate email for brand {self.brand}: {email}")
            return 'skipped'

        # Handle domain authority
        if 'domain_authority' in data:
            try:
                da = int(data['domain_authority'])
                data['domain_authority'] = max(0, min(100, da))  # Clamp 0-100
            except (ValueError, TypeError):
                data.pop('domain_authority', None)

        # Handle tags
        if 'tags' in data:
            tags_str = data['tags']
            if isinstance(tags_str, str):
                data['tags'] = [t.strip() for t in tags_str.split(',') if t.strip()]

        # Handle website - ensure it has protocol
        if 'website' in data:
            website = data['website']
            if website and not website.startswith(('http://', 'https://')):
                data['website'] = f"https://{website}"

        # Create contact
        with transaction.atomic():
            contact = Contact.objects.create(
                brand=self.brand,
                email=data['email'],
                name=data.get('name', email.split('@')[0]),
                company=data.get('company', ''),
                website=data.get('website', ''),
                domain_authority=data.get('domain_authority'),
                contact_type=self.contact_type,
                source=self.source,
                tags=data.get('tags', []),
                notes=data.get('notes', ''),
            )

            # Create deal if configured
            if self.create_deals and self.pipeline:
                first_stage = self.pipeline.stages.order_by('order').first()
                if first_stage:
                    Deal.objects.create(
                        contact=contact,
                        pipeline=self.pipeline,
                        current_stage=first_stage,
                        status='active',
                        next_action_date=timezone.now(),
                    )

        return 'imported'


def preview_file(file_content: bytes, filename: str = 'file.csv', max_rows: int = 5) -> Dict:
    """
    Preview file contents (CSV or Excel) and detected columns.

    Args:
        file_content: Raw file bytes
        filename: Filename (used to detect format)
        max_rows: Maximum rows to preview

    Returns:
        Dict with headers, mapped_columns, and sample_rows
    """
    try:
        headers, all_rows = parse_file_to_rows(file_content, filename)

        # Map columns
        mapped = {}
        for col in headers:
            col_lower = col.lower().strip()
            if col_lower in COLUMN_MAPPINGS:
                mapped[col] = COLUMN_MAPPINGS[col_lower]
            else:
                for key, field in COLUMN_MAPPINGS.items():
                    if key in col_lower or col_lower in key:
                        mapped[col] = field
                        break
                if col not in mapped:
                    mapped[col] = None  # Unmapped

        # Get sample rows
        sample_rows = all_rows[:max_rows]

        return {
            'success': True,
            'headers': headers,
            'mapped_columns': mapped,
            'sample_rows': sample_rows,
            'total_rows': len(all_rows),
            'has_email': 'email' in mapped.values(),
        }

    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }


# Backward compatibility aliases
CSVContactImporter = ContactImporter
preview_csv = preview_file
